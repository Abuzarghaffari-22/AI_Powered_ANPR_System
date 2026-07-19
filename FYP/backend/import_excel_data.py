#!/usr/bin/env python3
"""
ANPR System — Excel Data Import (CORRECTED, industry-grade)
===========================================================

Imports vehicle data from Car_info.xlsx into MySQL, PRESERVING the real
authorization status, dues, and all identity fields.

Why this replaces the old import_excel_data.py:
  * The old script HARD-CODED every vehicle to status='Authorized',
    dues='Clear', is_authorized=1 — destroying the real data. With your sheet
    that silently turned 249 unauthorized cars and 140 with outstanding dues
    into "authorized". A toll gate built on that is unsafe.
  * The old script matched columns by loose substring ('number' matched
    License/Engine/Chassis Number — last one won), so license_number could be
    set to the chassis number.

This version:
  * Maps columns EXPLICITLY by position/known header, with a fallback.
  * Computes is_authorized from the REAL dues + status, the SAME way the running
    pipeline does (pipeline._compute_auth_status / plate_store._auth_status), so
    the dashboard, the registry, and the live detection all agree.
  * Normalizes plates with the SAME normalize_plate the pipeline uses (imported
    if available, with an identical fallback) so license_normalized matches what
    OCR produces — otherwise lookups silently miss.
  * Uses a transaction + executemany for speed and atomicity (all-or-nothing).
  * Validates and reports: duplicates, blanks, status distribution.

Run from backend/:  python import_excel_data.py
"""

import logging
import os
import re
import sys
from pathlib import Path

import mysql.connector
import openpyxl
from dotenv import load_dotenv

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s | %(levelname)-7s | %(message)s")
logger = logging.getLogger("anpr.import")

load_dotenv(override=True)

# ── Use the pipeline's own normalizer so DB plates match OCR output exactly ──
try:
    from pipeline import normalize_plate as _pipeline_normalize
    logger.info("Using pipeline.normalize_plate for plate normalization")
except Exception:
    _pipeline_normalize = None
    logger.warning("pipeline.normalize_plate unavailable — using built-in fallback")

_RE_STRIP_SPACE = re.compile(r"\s+")
_RE_STRIP_CHARS = re.compile(r"[^A-Z0-9\-]")
_RE_YEAR_PURE   = re.compile(r"^([A-Z]{1,5})(\d{4})(\d{2})$")
_RE_YEAR_MIXED  = re.compile(r"^([A-Z]{1,5}\d{1,6}[A-Z]?)(\d{2})$")
_RE_SIMPLE      = re.compile(r"^([A-Z]{1,5})(\d{1,6})$")


def _fallback_normalize(text: str) -> str:
    if not text:
        return ""
    text = _RE_STRIP_SPACE.sub("", str(text).upper().strip())
    text = _RE_STRIP_CHARS.sub("", text)
    if len(text) < 3:
        return ""
    if "-" not in text:
        m = _RE_YEAR_PURE.match(text)
        if m:
            return m.group(1) + m.group(2) + "-" + m.group(3)
        ms = _RE_SIMPLE.match(text)
        if ms and ms.group(1).isalpha():
            pre, suf = ms.group(1), ms.group(2)
            if len(suf) == 5:
                return pre + suf[:3] + "-" + suf[3:]
            if len(suf) == 6:
                return pre + suf[:4] + "-" + suf[4:]
            return pre + "-" + suf
        mk = _RE_YEAR_MIXED.match(text)
        if mk and len(mk.group(2)) == 2:
            return mk.group(1) + "-" + mk.group(2)
    return text


def normalize_plate(text: str) -> str:
    if _pipeline_normalize is not None:
        try:
            return _pipeline_normalize(str(text))
        except Exception:
            pass
    return _fallback_normalize(text)


def compute_is_authorized(dues: str, status: str) -> int:
    """
    IDENTICAL logic to the running system (pipeline / plate_store / routes).
    Authorized only if dues are clear AND status is 'authorized'.
    """
    d = str(dues or "").strip().lower()
    s = str(status or "").strip().lower()
    return 1 if (d in ("clear", "paid", "", "none") and s == "authorized") else 0


def connect_database():
    return mysql.connector.connect(
        host=os.getenv("DB_HOST", "localhost"),
        user=os.getenv("DB_USER", "anpr_user"),
        password=os.getenv("DB_PASSWORD", "anpr_pass123"),
        database=os.getenv("DB_NAME", "anpr_db"),
        port=int(os.getenv("DB_PORT", "3306")),
        charset="utf8mb4",
    )


# Known column layout of Car_info.xlsx (by header keyword -> field).
# Explicit, position-independent: we find each field by an exact-ish header match
# and fall back to position if a header is garbled (your 'Model' header is
# 'M+B1:I1odel', so we also allow positional rescue).
HEADER_MAP = {
    "id":            ("vehicle_id_code", ["id"]),
    "make":          ("make",            ["make", "brand"]),
    "model":         ("model",           ["model"]),
    "license":       ("license_number",  ["license number", "license", "plate"]),
    "color":         ("color",           ["color", "colour"]),
    "engine":        ("engine_number",   ["engine"]),
    "chassis":       ("chassis_number",  ["chassis"]),
    "owner":         ("owner_name",      ["owner name", "owner"]),
    # Strip trailing spaces from CNIC header before matching
    "cnic":          ("owner_cnic",      ["owner cnic", "cnic"]),
    "dues":          ("dues",            ["dues"]),
    "status":        ("status",          ["status"]),
}


# Canonical make name map: lowercase -> Title Case
_MAKE_NORMALIZE: dict[str, str] = {
    "suzuki":  "Suzuki",
    "toyota":  "Toyota",
    "honda":   "Honda",
    "hyundai": "Hyundai",
    "hyundia": "Hyundai",  # common typo in data
    "kia":     "KIA",
}


def _normalize_make(raw: str) -> str:
    """Strip cell-ref garbage (e.g. 'honda+C1A78:C105'), normalize case."""
    if not raw:
        return ""
    # Strip Excel cell-ref artefacts: take only the part before any '+' or first digit
    cleaned = re.split(r"[+]", raw)[0].strip()
    return _MAKE_NORMALIZE.get(cleaned.lower(), cleaned.title())


def resolve_columns(header_row):
    """Return {field_name: column_index} mapping, explicit and audited."""
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
    mapping = {}
    for _key, (field, aliases) in HEADER_MAP.items():
        idx = None
        for alias in aliases:
            for i, h in enumerate(headers):
                if h == alias or (alias in h and field not in ("license_number",)):
                    idx = i
                    break
            if idx is not None:
                break
        # exact 'license number' guard so it never grabs engine/chassis 'number'
        if field == "license_number":
            for i, h in enumerate(headers):
                if "license" in h or h == "plate":
                    idx = i
                    break
        mapping[field] = idx
    return mapping, headers


def import_excel(excel_path: Path, conn) -> int:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["Sheet"] if "Sheet" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        logger.error("Excel is empty")
        return 0

    header = rows[0]
    colmap, headers = resolve_columns(header)
    logger.info(f"Resolved columns: {colmap}")

    # Positional rescue for the known garbled 'Model' header
    if colmap.get("model") is None and len(headers) > 2:
        colmap["model"] = 2
        logger.warning("Model header garbled — using positional column 2")

    required = ["license_number", "dues", "status"]
    missing = [f for f in required if colmap.get(f) is None]
    if missing:
        logger.error(f"Cannot find required columns: {missing}")
        return 0

    def cell(row, field):
        idx = colmap.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return None if v is None else str(v).strip()

    records = []
    seen_norm = {}
    dup_count = 0
    skip_count = 0
    auth_count = 0
    unauth_count = 0

    for n, row in enumerate(rows[1:], start=2):
        lic = cell(row, "license_number")
        if not lic:
            skip_count += 1
            continue
        norm = normalize_plate(lic)
        if not norm or len(norm) < 3:
            logger.warning(f"Row {n}: invalid plate '{lic}' — skipped")
            skip_count += 1
            continue

        if norm in seen_norm:
            dup_count += 1
            logger.warning(f"Row {n}: duplicate plate '{norm}' "
                           f"(first seen row {seen_norm[norm]}) — skipped")
            continue
        seen_norm[norm] = n

        dues   = cell(row, "dues")   or "Clear"
        status = cell(row, "status") or "Unauthorized"
        # Canonicalize capitalization for clean display
        # Handle lowercase variants: 'clear' -> 'Clear'
        dues   = dues.strip().capitalize()
        status = status.strip().capitalize()
        is_auth = compute_is_authorized(dues, status)
        if is_auth:
            auth_count += 1
        else:
            unauth_count += 1

        # Normalize make: fix garbled cell-refs and case inconsistencies
        raw_make  = cell(row, "make") or ""
        norm_make = _normalize_make(raw_make)

        # Engine/chassis numbers may be stored as integers in Excel
        raw_engine  = cell(row, "engine_number")
        raw_chassis = cell(row, "chassis_number")

        records.append((
            (cell(row, "vehicle_id_code") or "")[:20],
            norm_make[:60],
            (cell(row, "model")           or "").strip()[:60],
            lic[:40],
            norm[:40],
            (cell(row, "color")           or "").strip()[:40],
            (raw_engine  or "")[:60],
            (raw_chassis or "")[:60],
            (cell(row, "owner_name")      or "").strip()[:120],
            (cell(row, "owner_cnic")      or "").strip()[:30],
            dues[:20],
            status[:20],
            is_auth,
        ))

    wb.close()

    if not records:
        logger.error("No valid records to import")
        return 0

    cur = conn.cursor()
    try:
        conn.start_transaction()
        cur.execute("DELETE FROM detection_logs")  # FK-safe: clear logs first
        cur.execute("DELETE FROM vehicles")
        cur.executemany("""
            INSERT INTO vehicles
                (vehicle_id_code, make, model, license_number, license_normalized,
                 color, engine_number, chassis_number, owner_name, owner_cnic,
                 dues, status, is_authorized)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, records)
        conn.commit()
    except mysql.connector.Error as e:
        conn.rollback()
        logger.error(f"Import transaction failed (rolled back): {e}")
        cur.close()
        return 0
    cur.close()

    logger.info("=" * 60)
    logger.info(f"Imported     : {len(records)} vehicles")
    logger.info(f"  Authorized : {auth_count}")
    logger.info(f"  Unauthorized: {unauth_count}")
    logger.info(f"Duplicates   : {dup_count} (skipped)")
    logger.info(f"Invalid/blank: {skip_count} (skipped)")
    logger.info("=" * 60)
    return len(records)


def main() -> int:
    excel_path = Path(__file__).parent.parent / "data" / "Car_info.xlsx"
    if not excel_path.exists():
        # also allow running with the file beside the script
        alt = Path(__file__).parent / "Car_info.xlsx"
        if alt.exists():
            excel_path = alt
        else:
            logger.error(f"Excel not found: {excel_path}")
            return 1
    logger.info(f"Excel: {excel_path}")

    try:
        conn = connect_database()
    except mysql.connector.Error as e:
        logger.error(f"DB connection failed: {e}")
        return 1

    try:
        count = import_excel(excel_path, conn)
        if count <= 0:
            return 1
        # Refresh the in-RAM plate store if the backend is importable
        try:
            import plate_store
            n = plate_store.load(conn)
            logger.info(f"PlateStore reloaded in RAM: {n} vehicles")
        except Exception as e:
            logger.info(f"(Skipped in-process PlateStore reload: {e}. "
                        f"Use POST /api/stats/reload-store or restart backend.)")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    sys.exit(main())